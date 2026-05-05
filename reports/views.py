from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count, Sum, Avg, Q
from datetime import date
import json
import calendar

from accounts.decorators import staff_required


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _money_sum(qs):
    return float(qs.aggregate(t=Sum("estimated_cost"))["t"] or 0)


def _prev_next_month(year, month):
    if month == 1:
        prev = f"{year - 1}-12"
    else:
        prev = f"{year}-{month - 1:02d}"

    if month == 12:
        nxt = f"{year + 1}-01"
    else:
        nxt = f"{year}-{month + 1:02d}"

    return prev, nxt


def _week_labels(year, month):
    _, last_day = calendar.monthrange(year, month)

    labels = []
    d = 1

    while d <= last_day:
        end = min(d + 6, last_day)
        labels.append(f"T{d}-{end}/{month}")
        d += 7

    return labels[:4]


def _week_buckets(qs, year, month, date_field="created_at"):
    _, last_day = calendar.monthrange(year, month)

    cuts = [1, 8, 15, 22, last_day + 1]
    counts = []

    for i in range(4):
        s = date(year, month, cuts[i])
        e = date(year, month, min(cuts[i + 1] - 1, last_day))

        counts.append(
            qs.filter(
                **{
                    f"{date_field}__date__gte": s,
                    f"{date_field}__date__lte": e,
                }
            ).count()
        )

    return counts


def _week_costs(qs, year, month):
    _, last_day = calendar.monthrange(year, month)

    cuts = [1, 8, 15, 22, last_day + 1]
    costs = []

    for i in range(4):
        s = date(year, month, cuts[i])
        e = date(year, month, min(cuts[i + 1] - 1, last_day))

        cost = _money_sum(
            qs.filter(
                created_at__date__gte=s,
                created_at__date__lte=e,
            )
        )

        costs.append(round(cost / 1_000_000, 2))

    return costs


def _type_display(value):
    type_map = {
        "road_damage": "Hư hỏng đường",
        "pothole": "Ổ gà",
        "light_issue": "Đèn giao thông",
        "flood": "Ngập lụt",
        "other": "Khác",
    }
    return type_map.get(value, value or "Không rõ")


def _get_area_name(row):
    return (
        row.get("road__name")
        or row.get("infrastructure__name")
        or "Không rõ"
    )


# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------

@login_required
@staff_required
def report_summary(request):
    from maintenance.models import MaintenanceRequest
    from infrastructure.models import Road, TrafficLight

    now = timezone.now()
    month_start = now.date().replace(day=1)

    qs_all = MaintenanceRequest.objects.all()
    qs_month = qs_all.filter(created_at__date__gte=month_start)

    incidents_by_type = list(
        qs_all.values("incident_type").annotate(count=Count("id"))
    )

    incidents_by_status = list(
        qs_all.values("status").annotate(count=Count("id"))
    )

    total_cost = _money_sum(qs_all)

    avg_rating = qs_all.filter(
        citizen_rating__isnull=False
    ).aggregate(avg=Avg("citizen_rating"))["avg"]

    six_months = []
    base = now.date().replace(day=1)

    for i in range(5, -1, -1):
        d = base
        for _ in range(i):
            if d.month == 1:
                d = d.replace(year=d.year - 1, month=12)
            else:
                d = d.replace(month=d.month - 1)

        count = qs_all.filter(
            created_at__year=d.year,
            created_at__month=d.month
        ).count()

        six_months.append({
            "month": f"{d.month}/{d.year}",
            "count": count,
        })

    stats = {
        "total_incidents": qs_all.count(),
        "monthly_incidents": qs_month.count(),
        "completed_incidents": qs_all.filter(status="completed").count(),
        "pending_incidents": qs_all.filter(status="pending").count(),
        "total_roads": Road.objects.count(),
        "total_lights": TrafficLight.objects.count(),
        "total_cost": total_cost,
        "avg_rating": round(avg_rating, 1) if avg_rating else 0,
        "incidents_by_type": incidents_by_type,
        "incidents_by_status": incidents_by_status,
        "six_months": six_months,
    }

    type_labels = [_type_display(x["incident_type"]) for x in incidents_by_type]
    type_counts = [x["count"] for x in incidents_by_type]

    status_map = {
        "pending": "Chờ xử lý",
        "in_progress": "Đang xử lý",
        "completed": "Hoàn thành",
        "rejected": "Từ chối",
    }

    status_labels = [
        status_map.get(x["status"], x["status"] or "Không rõ")
        for x in incidents_by_status
    ]

    status_counts = [x["count"] for x in incidents_by_status]

    return render(request, "reports/summary.html", {
        "stats": stats,
        "type_labels": json.dumps(type_labels, ensure_ascii=False),
        "type_counts": json.dumps(type_counts),
        "status_labels": json.dumps(status_labels, ensure_ascii=False),
        "status_counts": json.dumps(status_counts),
        "six_months_labels": json.dumps([x["month"] for x in six_months]),
        "six_months_counts": json.dumps([x["count"] for x in six_months]),
        "page_title": "Báo cáo tổng hợp",
    })


# ---------------------------------------------------------------------------
# Monthly
# ---------------------------------------------------------------------------

@login_required
@staff_required
def monthly_report(request):
    from maintenance.models import MaintenanceRequest

    now = timezone.now()
    month_param = request.GET.get("month", "")

    if "-" in month_param:
        try:
            year = int(month_param.split("-")[0])
            month = int(month_param.split("-")[1])
        except (ValueError, IndexError):
            year = now.year
            month = now.month
    else:
        year = int(request.GET.get("year", now.year))
        month = int(request.GET.get("month", now.month))

    selected_month = f"{year}-{month:02d}"
    prev_month, next_month = _prev_next_month(year, month)

    qs = MaintenanceRequest.objects.filter(
        created_at__year=year,
        created_at__month=month,
    )

    total = qs.count()
    resolved = qs.filter(status="completed").count()
    in_progress = qs.filter(status="in_progress").count()
    pending = qs.filter(status="pending").count()
    rejected = qs.filter(status="rejected").count()

    if month == 1:
        prev_qs = MaintenanceRequest.objects.filter(
            created_at__year=year - 1,
            created_at__month=12,
        )
    else:
        prev_qs = MaintenanceRequest.objects.filter(
            created_at__year=year,
            created_at__month=month - 1,
        )

    prev_total = prev_qs.count() or 1
    requests_vs_last = round((total - prev_total) / prev_total * 100)
    resolve_rate = round(resolved / total * 100) if total else 0

    completed_with_time = qs.filter(
        status="completed",
        completed_at__isnull=False,
    )

    if completed_with_time.exists():
        total_days = 0
        valid_count = 0

        for r in completed_with_time:
            if r.completed_at and r.created_at:
                total_days += (r.completed_at.date() - r.created_at.date()).days
                valid_count += 1

        avg_days = round(total_days / valid_count, 1) if valid_count else 0
    else:
        avg_days = 0

    total_cost = _money_sum(qs)
    avg_cost = round(total_cost / total / 1000) if total else 0

    kpi = {
        "total_requests": total,
        "requests_vs_last": requests_vs_last,
        "resolved": resolved,
        "resolve_rate": resolve_rate,
        "avg_days": avg_days,
        "total_cost": round(total_cost / 1_000_000, 1),
        "avg_cost": avg_cost,
        "in_progress": in_progress,
        "pending": pending,
        "rejected": rejected,
    }

    week_labels_list = _week_labels(year, month)
    weekly_requests = _week_buckets(qs, year, month)
    weekly_resolved = _week_buckets(qs.filter(status="completed"), year, month)
    weekly_costs = _week_costs(qs, year, month)

    by_type = list(
        qs.values("incident_type")
        .annotate(count=Count("id"))
        .order_by("-count")
    )

    type_labels = [_type_display(x["incident_type"]) for x in by_type]
    type_counts = [x["count"] for x in by_type]

    top_area_qs = (
        qs.values(
            "road__name",
            "infrastructure__name",
        )
        .annotate(
            count=Count("id"),
            resolved_count=Count("id", filter=Q(status="completed")),
        )
        .order_by("-count")[:5]
    )

    top_districts = []

    for d in top_area_qs:
        name = _get_area_name(d)
        count = d["count"]
        resolved_count = d["resolved_count"]

        top_districts.append({
            "name": name,
            "count": count,
            "resolved": resolved_count,
            "rate": round(resolved_count / count * 100) if count else 0,
        })

    staff_qs = (
        qs.filter(
            status="completed",
            assigned_to__isnull=False,
        )
        .values(
            "assigned_to__id",
            "assigned_to__first_name",
            "assigned_to__last_name",
            "assigned_to__username",
        )
        .annotate(completed=Count("id"))
        .order_by("-completed")[:5]
    )

    top_staff = []

    for s in staff_qs:
        first_name = s.get("assigned_to__first_name") or ""
        last_name = s.get("assigned_to__last_name") or ""
        username = s.get("assigned_to__username") or ""

        name = f"{last_name} {first_name}".strip() or username or "N/A"

        top_staff.append({
            "name": name,
            "completed": s["completed"],
            "avg_days": 4,
        })

    return render(request, "reports/monthly.html", {
        "kpi": kpi,
        "selected_month": selected_month,
        "prev_month": prev_month,
        "next_month": next_month,

        "week_labels": json.dumps(week_labels_list, ensure_ascii=False),
        "weekly_requests": json.dumps(weekly_requests),
        "weekly_resolved": json.dumps(weekly_resolved),
        "weekly_costs": json.dumps(weekly_costs),

        "type_labels": json.dumps(type_labels, ensure_ascii=False),
        "type_counts": json.dumps(type_counts),

        "top_districts": top_districts,
        "top_staff": top_staff,

        "page_title": f"Báo cáo tháng {month}/{year}",
    })


# ---------------------------------------------------------------------------
# Yearly
# ---------------------------------------------------------------------------

@login_required
@staff_required
def yearly_report(request):
    from maintenance.models import MaintenanceRequest

    now = timezone.now()

    try:
        year = int(request.GET.get("year", now.year))
    except ValueError:
        year = now.year

    year_list = list(range(now.year, now.year - 5, -1))

    month_names = [
        "T1", "T2", "T3", "T4", "T5", "T6",
        "T7", "T8", "T9", "T10", "T11", "T12"
    ]

    monthly_data = []
    total_year = 0
    total_cost_year = 0.0

    for m in range(1, 13):
        qs_month = MaintenanceRequest.objects.filter(
            created_at__year=year,
            created_at__month=m,
        )

        count = qs_month.count()
        completed = qs_month.filter(status="completed").count()
        cost = _money_sum(qs_month)

        total_year += count
        total_cost_year += cost

        monthly_data.append({
            "month": m,
            "label": month_names[m - 1],
            "count": count,
            "completed": completed,
            "cost": round(cost / 1_000_000, 1),
            "rate": round(completed / count * 100) if count else 0,
        })

    qs_year = MaintenanceRequest.objects.filter(created_at__year=year)

    completed_year = qs_year.filter(status="completed").count()
    pending_year = qs_year.filter(status="pending").count()
    in_progress_year = qs_year.filter(status="in_progress").count()

    prev_year_total = MaintenanceRequest.objects.filter(
        created_at__year=year - 1
    ).count() or 1

    vs_prev = round((total_year - prev_year_total) / prev_year_total * 100)

    yearly_stats = {
        "total": total_year,
        "completed": completed_year,
        "pending": pending_year,
        "in_progress": in_progress_year,
        "resolve_rate": round(completed_year / total_year * 100) if total_year else 0,
        "total_cost": round(total_cost_year / 1_000_000, 1),
        "vs_prev": vs_prev,
    }

    by_type = list(
        qs_year.values("incident_type")
        .annotate(count=Count("id"))
        .order_by("-count")
    )

    type_labels = [_type_display(x["incident_type"]) for x in by_type]
    type_counts = [x["count"] for x in by_type]

    top_area_qs = (
        qs_year.values(
            "road__name",
            "infrastructure__name",
        )
        .annotate(count=Count("id"))
        .order_by("-count")[:5]
    )

    top_districts = []

    for d in top_area_qs:
        top_districts.append({
            "name": _get_area_name(d),
            "count": d["count"],
        })

    counts_by_month = [d["count"] for d in monthly_data]
    completed_by_month = [d["completed"] for d in monthly_data]
    costs_by_month = [d["cost"] for d in monthly_data]

    return render(request, "reports/yearly.html", {
        "year": year,
        "year_list": year_list,
        "monthly_data": monthly_data,
        "yearly_stats": yearly_stats,
        "top_districts": top_districts,

        "months": json.dumps(month_names, ensure_ascii=False),
        "data_months": json.dumps(counts_by_month),
        "data_completed": json.dumps(completed_by_month),
        "data_costs": json.dumps(costs_by_month),
        "type_labels": json.dumps(type_labels, ensure_ascii=False),
        "type_counts": json.dumps(type_counts),

        "page_title": f"Báo cáo năm {year}",
    })


# ---------------------------------------------------------------------------
# Export Excel
# ---------------------------------------------------------------------------

@login_required
@staff_required
def export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from maintenance.models import MaintenanceRequest

    month_param = request.GET.get("month", "")
    year_param = request.GET.get("year", "")

    if "-" in month_param:
        try:
            year = int(month_param.split("-")[0])
            month = int(month_param.split("-")[1])

            qs = MaintenanceRequest.objects.filter(
                created_at__year=year,
                created_at__month=month,
            )

            filename = f"bao_cao_{year}_{month:02d}.xlsx"
        except (ValueError, IndexError):
            qs = MaintenanceRequest.objects.all()
            filename = "bao_cao_su_co.xlsx"

    elif year_param:
        try:
            year = int(year_param)
            qs = MaintenanceRequest.objects.filter(created_at__year=year)
            filename = f"bao_cao_nam_{year}.xlsx"
        except ValueError:
            qs = MaintenanceRequest.objects.all()
            filename = "bao_cao_su_co.xlsx"

    else:
        qs = MaintenanceRequest.objects.all()
        filename = "bao_cao_su_co.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Báo cáo sự cố"

    header_fill = PatternFill("solid", fgColor="1976D2")
    header_font = Font(color="FFFFFF", bold=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "STT",
        "Tiêu đề",
        "Loại sự cố",
        "Mức độ ưu tiên",
        "Trạng thái",
        "Người báo",
        "Ngày tạo",
        "Ngày hoàn thành",
        "Chi phí dự kiến (VNĐ)",
    ]

    col_widths = [6, 35, 18, 16, 14, 22, 14, 16, 22]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 22

    status_map = {
        "pending": "Chờ xử lý",
        "in_progress": "Đang xử lý",
        "completed": "Hoàn thành",
        "rejected": "Từ chối",
    }

    for i, r in enumerate(qs.select_related("reported_by"), 1):
        row_data = [
            i,
            r.title,
            r.get_incident_type_display() if hasattr(r, "get_incident_type_display") else r.incident_type,
            r.get_priority_display() if hasattr(r, "get_priority_display") else r.priority,
            status_map.get(r.status, r.status),
            r.reported_by.get_full_name() if r.reported_by else "",
            r.created_at.strftime("%d/%m/%Y") if r.created_at else "",
            r.completed_at.strftime("%d/%m/%Y") if r.completed_at else "",
            float(getattr(r, "estimated_cost", 0) or 0),
        ]

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=ci, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            if ci == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")

            if ci == 9:
                cell.number_format = "#,##0"

        if i % 2 == 0:
            for ci in range(1, 10):
                ws.cell(row=i + 1, column=ci).fill = PatternFill(
                    "solid",
                    fgColor="F5F5F5",
                )

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    resp["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(resp)
    return resp


@login_required
@staff_required
def export_monthly_excel(request):
    return export_excel(request)


# ---------------------------------------------------------------------------
# Export PDF
# ---------------------------------------------------------------------------

@login_required
@staff_required
def export_pdf(request):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from maintenance.models import MaintenanceRequest

    now = timezone.now()

    month_param = request.GET.get("month", "")
    year_param = request.GET.get("year", "")

    if "-" in month_param:
        try:
            year = int(month_param.split("-")[0])
            month = int(month_param.split("-")[1])

            qs = MaintenanceRequest.objects.filter(
                created_at__year=year,
                created_at__month=month,
            )

            title_str = f"BAO CAO SU CO THANG {month:02d}/{year}"
        except (ValueError, IndexError):
            qs = MaintenanceRequest.objects.all()
            title_str = "BAO CAO SU CO HA TANG DO THI"

    elif year_param:
        try:
            year = int(year_param)

            qs = MaintenanceRequest.objects.filter(created_at__year=year)
            title_str = f"BAO CAO SU CO NAM {year}"
        except ValueError:
            qs = MaintenanceRequest.objects.all()
            title_str = "BAO CAO SU CO HA TANG DO THI"

    else:
        qs = MaintenanceRequest.objects.all()
        title_str = "BAO CAO SU CO HA TANG DO THI"

    resp = HttpResponse(content_type="application/pdf")
    resp["Content-Disposition"] = 'attachment; filename="bao_cao.pdf"'

    p = rl_canvas.Canvas(resp, pagesize=A4)
    w, h = A4

    p.setFillColor(colors.HexColor("#1976D2"))
    p.rect(0, h - 3 * cm, w, 3 * cm, fill=True, stroke=False)

    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 16)
    p.drawCentredString(w / 2, h - 1.5 * cm, title_str)

    p.setFont("Helvetica", 11)
    p.drawCentredString(
        w / 2,
        h - 2.3 * cm,
        f"Ngay xuat: {now.strftime('%d/%m/%Y %H:%M')}",
    )

    total = qs.count()
    completed = qs.filter(status="completed").count()
    pending = qs.filter(status="pending").count()
    in_progress = qs.filter(status="in_progress").count()
    total_cost = _money_sum(qs)

    p.setFillColor(colors.black)
    p.setFont("Helvetica-Bold", 13)
    p.drawString(2 * cm, h - 4 * cm, "THONG KE TONG QUAN")

    rows = [
        ("Tong so su co:", str(total)),
        ("Da hoan thanh:", str(completed)),
        ("Dang xu ly:", str(in_progress)),
        ("Cho xu ly:", str(pending)),
        ("Ty le hoan thanh:", f"{round(completed / total * 100) if total else 0}%"),
        ("Tong chi phi du kien (trieu dong):", f"{round(total_cost / 1_000_000, 1):,.1f}M"),
    ]

    y = h - 5 * cm

    for label, val in rows:
        p.setFont("Helvetica-Bold", 11)
        p.drawString(2 * cm, y, label)

        p.setFont("Helvetica", 11)
        p.drawString(10 * cm, y, val)

        y -= 0.8 * cm

    p.setFont("Helvetica-Bold", 13)
    p.drawString(2 * cm, y - 0.5 * cm, "CHI TIET SU CO")

    y -= 1.5 * cm

    p.setFont("Helvetica-Bold", 9)
    p.setFillColor(colors.HexColor("#1976D2"))
    p.rect(2 * cm, y, w - 4 * cm, 0.6 * cm, fill=True, stroke=False)

    p.setFillColor(colors.white)
    p.drawString(2.2 * cm, y + 0.15 * cm, "STT")
    p.drawString(3.5 * cm, y + 0.15 * cm, "Tieu de")
    p.drawString(11 * cm, y + 0.15 * cm, "Trang thai")
    p.drawString(15 * cm, y + 0.15 * cm, "Chi phi")

    y -= 0.6 * cm

    p.setFont("Helvetica", 9)
    p.setFillColor(colors.black)

    for i, r in enumerate(qs[:25], 1):
        if y < 2 * cm:
            p.showPage()
            y = h - 2 * cm

        if i % 2 == 0:
            p.setFillColor(colors.HexColor("#F5F5F5"))
            p.rect(2 * cm, y, w - 4 * cm, 0.55 * cm, fill=True, stroke=False)
            p.setFillColor(colors.black)

        title = r.title or ""
        title_short = title[:45] + "..." if len(title) > 45 else title

        p.drawString(2.2 * cm, y + 0.12 * cm, str(i))
        p.drawString(3.5 * cm, y + 0.12 * cm, title_short)
        p.drawString(11 * cm, y + 0.12 * cm, r.status or "")
        p.drawString(
            15 * cm,
            y + 0.12 * cm,
            f"{float(getattr(r, 'estimated_cost', 0) or 0):,.0f}",
        )

        y -= 0.55 * cm

    p.showPage()
    p.save()

    return resp


@login_required
@staff_required
def export_monthly_pdf(request):
    return export_pdf(request)